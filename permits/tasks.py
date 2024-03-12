import copy
import logging
from datetime import datetime
import calendar
import io

from django.contrib.auth.models import Group
from django.db.models import Sum, F
from django.urls import reverse_lazy

from celery import shared_task

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment

from users.models import (
    User,
    Admin,
    Client,
    Notification
)

from payments.models import PaymentOrder, PaymentOrderItem

from .models import (
    PermitApplication,
    Status,
    Permit,
    LocalTransportPermit,
    TransportEntry
)
from .emails import (
    SubmittedApplicationEmailView,
    UnsubmittedApplicationEmailView,
    AcceptedApplicationEmailView,
    ReturnedApplicationEmailView,
    ScheduledInspectionEmailView,
    AssignedScheduledInspectionEmailView,
    SignedInspectionEmailView,
    PermitCreatedEmailView,
    PermitSignedEmailView,
    PermitReleasedEmailView,
    PermitValidatedEmailView,
    PermitExpiredEmailView,
    ReportsEmailView
)


logger = logging.getLogger(__name__)

center = Alignment(horizontal='center', vertical='center',
                   wrapText=True)
top = Alignment(horizontal='left', vertical='top',
                wrapText=True)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"))
header_style = NamedStyle(name='header_style')
header_style.font = Font(bold=True)
header_style.alignment = center


def get_admins_who_can_receive_emails():
    admins = Admin.objects.filter(
        is_active=True, is_initial_password_changed=True)
    return admins


def get_permit_signatories_who_can_receive_emails():
    group = Group.objects.get(name='Permit Signatory')
    signatories = User.objects.filter(
        groups=group, is_active=True, is_initial_password_changed=True)
    return signatories


@shared_task
def notify_admins_about_submitted_application(application_id):
    application: PermitApplication = PermitApplication.objects.get(
        id=application_id)
    admins = get_admins_who_can_receive_emails()
    for admin in admins:
        SubmittedApplicationEmailView(admin, application).send()


@shared_task
def notify_admins_about_unsubmitted_application(application_id):
    application: PermitApplication = PermitApplication.objects.get(
        id=application_id)
    admins = get_admins_who_can_receive_emails()
    for admin in admins:
        UnsubmittedApplicationEmailView(admin, application).send()


@shared_task
def notify_client_about_accepted_application(application_id):
    application: PermitApplication = PermitApplication.objects.get(
        id=application_id)
    AcceptedApplicationEmailView(application.client, application).send()


@shared_task
def notify_client_about_returned_application(application_id):
    application: PermitApplication = PermitApplication.objects.get(
        id=application_id)
    ReturnedApplicationEmailView(application.client, application).send()


@shared_task
def notify_client_and_officer_about_scheduled_inspection(application_id):
    application: PermitApplication = PermitApplication.objects.get(
        id=application_id)
    ScheduledInspectionEmailView(
        application.client, application).send()


@shared_task
def notify_admins_about_signed_inspection(application_id):
    application: PermitApplication = PermitApplication.objects.get(
        id=application_id)
    admins = get_admins_who_can_receive_emails()
    for admin in admins:
        if admin.id != application.inspection.signatures.first().person.id:
            SignedInspectionEmailView(
                admin, application).send()


@shared_task
def notify_signatories_about_created_permit(permit_id):
    permit: Permit = Permit.objects.get(id=permit_id).subclass

    notifiable_permits = [LocalTransportPermit.__name__]
    if permit.type not in notifiable_permits:
        return

    signatories = get_permit_signatories_who_can_receive_emails()
    for signatory in signatories:
        PermitCreatedEmailView(signatory, permit).send()


@shared_task
def notify_admins_about_signed_permit(permit_id):
    permit: Permit = Permit.objects.get(id=permit_id).subclass
    admins = get_admins_who_can_receive_emails()
    for user in admins:
        PermitSignedEmailView(user, permit).send()


@shared_task
def notify_client_and_admins_about_released_permit(permit_id):
    permit: Permit = Permit.objects.get(id=permit_id).subclass
    users = list(get_admins_who_can_receive_emails())
    users.append(permit.client)
    for user in users:
        PermitReleasedEmailView(user, permit).send()


@shared_task
def notify_client_and_admins_about_validated_permit(permit_id):
    permit: Permit = Permit.objects.get(id=permit_id).subclass
    users = list(get_admins_who_can_receive_emails())
    users.append(permit.client)
    for user in users:
        PermitValidatedEmailView(user, permit).send()


@shared_task
def check_permit_validity():
    logger.info('Checking for permits to expire...')

    permits_to_expire = Permit.objects.filter(
        status__in=[Status.RELEASED],
        valid_until__lt=datetime.now().date())
    admins = list(get_admins_who_can_receive_emails())

    for permit in permits_to_expire:
        permit.status = Status.EXPIRED
        permit.save()
        logger.info('Permit %s has expired already.', permit.permit_no)

        # Notify the users
        users = copy.deepcopy(admins)
        users.append(permit.client)
        for user in users:
            PermitExpiredEmailView(user, permit.subclass).send()

        url = reverse_lazy('permit_detail', args=[permit.id])
        message = f'''
        Your permit <a href="{url}">{permit.permit_no}</a> has expired.
        '''
        Notification.objects.create(
            user=permit.client, message=message)

    logger.info('Done expiring permits.')


@shared_task
def generate_reports(year, quarter, user_id):
    quarters = [(1, 2, 3), (4, 5, 6), (7, 8, 9), (10, 11, 12)]
    months = quarters[int(quarter)-1]
    first_month, second_month, third_month = months

    # Whole quarter dates
    from_date = f'{year}-{first_month:02d}-01'
    to_date = f'{year}-{third_month:02d}-{calendar.monthrange(year, third_month)[1]}'

    # Month dates
    first_from_date = f'{year}-{first_month:02d}-01'
    first_to_date = f'{year}-{first_month:02d}-{calendar.monthrange(year, first_month)[1]}'
    second_from_date = f'{year}-{second_month:02d}-01'
    second_to_date = f'{year}-{second_month:02d}-{calendar.monthrange(year, second_month)[1]}'
    third_from_date = f'{year}-{third_month:02d}-01'
    third_to_date = f'{year}-{third_month:02d}-{calendar.monthrange(year, third_month)[1]}'
    month_dates = [
        (first_from_date, first_to_date),
        (second_from_date, second_to_date),
        (third_from_date, third_to_date)
    ]

    # Get the clients
    client_ids = LocalTransportPermit.objects.values('client').distinct()
    client_ids = list(map(lambda i: i['client'], client_ids))
    clients = Client.objects.filter(
        id__in=client_ids).order_by('first_name')
    data = {
        'clients': {
            'list': []
        }
    }
    for client in clients:
        client_data = {'client': client,
                       'permits_issued_in_month': [],
                       'total_permits': 0, 'fees_collected': 0,
                       'total_species': 0}
        ltp_filters = {
            'client': client,
            'status__in': [Status.RELEASED, Status.USED],
            'created_at__gte': from_date,
            'created_at__lte': to_date
        }
        for month_from_date, month_to_date in month_dates:
            ltp_filters['transport_date__gte'] = month_from_date
            ltp_filters['transport_date__lte'] = month_to_date
            permits_count = LocalTransportPermit.objects.filter(
                **ltp_filters).count()
            client_data['permits_issued_in_month'].append(permits_count)
            client_data['total_permits'] += permits_count

        ltp_filters['transport_date__gte'] = from_date
        ltp_filters['transport_date__lte'] = to_date
        ltp_ids = LocalTransportPermit.objects.filter(
            **ltp_filters).values('id')
        payment_orders = PaymentOrder.objects.filter(
            permit__id__in=ltp_ids)
        fees_collected = PaymentOrderItem.objects \
            .filter(payment_order__in=payment_orders) \
            .aggregate(total=Sum(F('amount')))['total'] or 0
        client_data['fees_collected'] = float(fees_collected)

        total_species = TransportEntry.objects \
            .filter(ltp__id__in=ltp_ids) \
            .aggregate(total=Sum(F('quantity')))['total'] or 0
        client_data['total_species'] = int(total_species)

        data['clients']['list'].append(client_data)

    wb = Workbook()
    wb.add_named_style(header_style)

    clients_sheet = wb.active
    build_clients_report(clients_sheet, data, year, quarter, months)

    ltp_sheet = wb.create_sheet('Local Transport Permits')
    build_ltp_reports(ltp_sheet, data, from_date, to_date)

    excel_bytesio = io.BytesIO()
    wb.save(excel_bytesio)
    user = User.objects.get(id=user_id)
    filename = f'reports_{datetime.now().strftime("%Y-%m-%d_%H%M%S")}.xlsx'
    ReportsEmailView(user).send(
        attachment=excel_bytesio.getvalue(), filename=filename)


def build_clients_report(ws, data, year, quarter, months):
    ws.title = 'Clients'

    first_month = calendar.month_name[months[0]]
    second_month = calendar.month_name[months[1]]
    third_month = calendar.month_name[months[2]]
    title = ('Issuance of Wildlife Local Transport Permit for Quarter '
             f'{quarter} ({first_month}-{third_month}) C.Y. - {year}')
    title = ws.cell(row=1, column=1, value=title)
    title.style = header_style
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    permittee = ws.cell(row=2, column=1, value='PERMITTEE')
    permittee.style = header_style
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.column_dimensions['a'].width = 20

    address = ws.cell(row=2, column=2, value='ADDRESS')
    address.style = header_style
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.column_dimensions['b'].width = 20

    business_name = ws.cell(row=2, column=3, value='BUSINESS NAME')
    business_name.style = header_style
    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
    ws.column_dimensions['c'].width = 20

    num_permits = ws.cell(
        row=2, column=4, value=f'NUMBER OF PERMITS ISSUED FOR QUARTER {quarter} C.Y. {year}')
    num_permits.style = header_style
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=6)
    ws.column_dimensions['d'].width = 15
    ws.column_dimensions['e'].width = 15
    ws.column_dimensions['f'].width = 15

    first_month = ws.cell(row=3, column=4, value=first_month.upper())
    first_month.style = header_style

    second_month = ws.cell(row=3, column=5, value=second_month.upper())
    second_month.style = header_style

    third_month = ws.cell(row=3, column=6, value=third_month.upper())
    third_month.style = header_style

    grand_total = ws.cell(row=2, column=7, value='GRAND TOTAL')
    grand_total.style = header_style
    ws.merge_cells(start_row=2, start_column=7, end_row=3, end_column=7)

    fees = ws.cell(row=2, column=8, value='FEES COLLECTED (Php)')
    fees.style = header_style
    ws.merge_cells(start_row=2, start_column=8, end_row=3, end_column=8)
    ws.column_dimensions['h'].width = 15

    total_species = ws.cell(
        row=2, column=9, value='TOTAL NO. OF SPECIES TRANSPORTED')
    total_species.style = header_style
    ws.merge_cells(start_row=2, start_column=9, end_row=3, end_column=9)
    ws.column_dimensions['i'].width = 20

    total_first_month = 0
    total_second_month = 0
    total_third_month = 0
    total_fees = 0
    total_species = 0
    grand_total = 0
    row_num = 4
    for i in data['clients']['list']:
        ws.cell(row=row_num, column=1,
                value=f"{i['client'].first_name} {i['client'].last_name}")
        ws.cell(row=row_num, column=2, value=i['client'].address)
        ws.cell(row=row_num, column=3, value=i['client'].current_farm_name)
        first_month_total = i['permits_issued_in_month'][0]
        total_first_month += first_month_total
        ws.cell(row=row_num, column=4, value=first_month_total)
        second_month_total = i['permits_issued_in_month'][1]
        total_second_month += second_month_total
        ws.cell(row=row_num, column=5, value=second_month_total)
        third_month_total = i['permits_issued_in_month'][2]
        total_third_month += third_month_total
        ws.cell(row=row_num, column=6, value=third_month_total)
        ws.cell(row=row_num, column=7,
                value=i['total_permits'])
        grand_total += i['total_permits']
        ws.cell(row=row_num, column=8,
                value=i['fees_collected'])
        total_fees += i['fees_collected']
        ws.cell(row=row_num, column=9,
                value=i['total_species'])
        total_species += i['total_species']
        row_num += 1

    total = ws.cell(row=row_num, column=3, value='TOTAL')
    total.style = header_style
    ws.cell(row=row_num, column=4, value=total_first_month)
    ws.cell(row=row_num, column=5, value=total_second_month)
    ws.cell(row=row_num, column=6, value=total_third_month)
    ws.cell(row=row_num, column=7, value=grand_total)
    ws.cell(row=row_num, column=8, value=total_fees)
    ws.cell(row=row_num, column=9, value=total_species)

    for col in range(1, 10):
        for row in range(1, row_num+1):
            ws.cell(row=row, column=col).border = border


def build_ltp_reports(ws, data, from_date, to_date):
    for col in 'abcdefghijklmnopqr':
        ws.column_dimensions[col].width = 20
    ws.column_dimensions['c'].width = 25

    permit_holder = ws.cell(row=1, column=1, value='Permit Holder')
    permit_holder.style = header_style
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    holder_name = ws.cell(row=2, column=1, value='Name of Holder')
    holder_name.style = header_style

    address = ws.cell(row=2, column=2, value='Address of Holder')
    address.style = header_style

    permit_details = ws.cell(row=1, column=3, value='Permit Details')
    permit_details.style = header_style
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=7)

    permit_no = ws.cell(row=2, column=3, value='Permit No.')
    permit_no.style = header_style

    permit_no = ws.cell(row=2, column=4, value='Date of Application')
    permit_no.style = header_style

    issuance = ws.cell(row=2, column=5, value='Issuance Date')
    issuance.style = header_style

    expiry = ws.cell(row=2, column=6, value='Expiry Date')
    expiry.style = header_style

    fees = ws.cell(row=2, column=7, value='Fees Collected')
    fees.style = header_style

    species_transported = ws.cell(row=1, column=8, value='Species Transported')
    species_transported.style = header_style
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=14)

    cell = ws.cell(row=2, column=8, value='Species Type')
    cell.style = header_style

    cell = ws.cell(row=2, column=9, value='Common Name')
    cell.style = header_style

    cell = ws.cell(row=2, column=10, value='Scientific Name')
    cell.style = header_style

    cell = ws.cell(row=2, column=11, value='Type of Specimen')
    cell.style = header_style

    cell = ws.cell(row=2, column=12, value='Quantity')
    cell.style = header_style

    cell = ws.cell(row=2, column=13, value='Unit of Measurement')
    cell.style = header_style

    cell = ws.cell(row=2, column=14, value='Description')
    cell.style = header_style

    cell = ws.cell(row=1, column=15, value='Origin')
    cell.style = header_style
    ws.merge_cells(start_row=1, start_column=15, end_row=2, end_column=15)

    cell = ws.cell(row=1, column=16, value='Destination')
    cell.style = header_style
    ws.merge_cells(start_row=1, start_column=16, end_row=2, end_column=16)

    cell = ws.cell(row=1, column=17, value='Purpose of Transfer')
    cell.style = header_style
    ws.merge_cells(start_row=1, start_column=17, end_row=2, end_column=17)

    cell = ws.cell(row=1, column=18, value='Remarks')
    cell.style = header_style
    ws.merge_cells(start_row=1, start_column=18, end_row=2, end_column=18)

    client_row = 3
    last_row = 0
    for client in data['clients']['list']:
        ltps = LocalTransportPermit.objects \
            .filter(
                client=client['client'],
                status__in=[Status.RELEASED, Status.USED],
                created_at__gte=from_date,
                created_at__lte=to_date) \
            .order_by('-created_at')
        if ltps.count() == 0:
            continue

        cell = ws.cell(row=client_row, column=1,
                       value=f"{client['client'].first_name} {client['client'].last_name}")
        cell.alignment = top
        cell = ws.cell(row=client_row, column=2,
                       value=client['client'].address)
        cell.alignment = top

        total_transport = 0
        permit_row = client_row
        for ltp in ltps:
            cell = ws.cell(row=permit_row, column=3, value=ltp.permit_no)
            cell.alignment = top
            if ltp.application:
                cell = ws.cell(row=permit_row, column=4,
                               value=ltp.application.created_at.date())
                cell.alignment = top
            cell = ws.cell(row=permit_row, column=5, value=ltp.issued_date)
            cell.alignment = top
            cell = ws.cell(row=permit_row, column=6, value=ltp.valid_until)
            cell.alignment = top
            if ltp.payment_order:
                cell = ws.cell(row=permit_row, column=7,
                               value=ltp.payment_order.total)
                cell.alignment = top

            species_row = permit_row
            transports = ltp.species_to_transport.order_by(
                'sub_species__main_species', 'sub_species__common_name')
            total_transport += transports.count()
            row_span = permit_row+transports.count()-1
            ws.merge_cells(start_row=permit_row, start_column=3,
                           end_row=row_span, end_column=3)
            ws.merge_cells(start_row=permit_row, start_column=4,
                           end_row=row_span, end_column=4)
            ws.merge_cells(start_row=permit_row, start_column=5,
                           end_row=row_span, end_column=5)
            ws.merge_cells(start_row=permit_row, start_column=6,
                           end_row=row_span, end_column=6)
            ws.merge_cells(start_row=permit_row, start_column=7,
                           end_row=row_span, end_column=7)
            for transport in transports:
                cell = ws.cell(row=species_row, column=8,
                               value=transport.sub_species.main_species.name)
                cell = ws.cell(row=species_row, column=9,
                               value=transport.sub_species.common_name)
                cell = ws.cell(row=species_row, column=10,
                               value=transport.sub_species.scientific_name)
                cell = ws.cell(row=species_row, column=11,
                               value=transport.description)
                cell = ws.cell(row=species_row, column=12,
                               value=transport.quantity)
                cell = ws.cell(row=species_row, column=13,
                               value='Pieces')
                cell = ws.cell(row=species_row, column=14,
                               value=transport.description)
                cell = ws.cell(row=species_row, column=15,
                               value='Marinduque')
                cell = ws.cell(row=species_row, column=16,
                               value=ltp.transport_location)
                cell = ws.cell(row=species_row, column=17,
                               value='Export')
                if ltp.application:
                    remarks = ltp.application.remarks.last()
                    if remarks:
                        cell = ws.cell(row=permit_row, column=18,
                                       value=remarks.content)

                species_row += 1
                permit_row += 1
                last_row = species_row

        row_span = client_row+total_transport-1
        ws.merge_cells(start_row=client_row, start_column=1,
                       end_row=row_span, end_column=1)
        ws.merge_cells(start_row=client_row, start_column=2,
                       end_row=row_span, end_column=2)

        client_row += total_transport

    for col in range(1, 19):
        for row in range(1, last_row):
            ws.cell(row=row, column=col).border = border
